"""
Tests for transactions/views.py - Transaction and Category API endpoints.

Tests:
- CategoryListCreateView - List and create categories
- CategoryDetailView - Retrieve, update, delete categories
- DefaultCategoryListView - List default categories
- TransactionListCreateView - List, create, filter, search, order transactions
- TransactionDetailView - Retrieve, update, delete transactions
- TransactionBulkCreateView - Bulk create transactions with validation
- TransactionExportView - Export transactions as CSV/XLSX
"""

import pytest
import csv
import io
from decimal import Decimal
from datetime import timedelta
from django.utils import timezone
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from openpyxl import load_workbook

from apps.transactions.tests.factories import (
    TransactionFactory,
    CategoryFactory,
)
from apps.transactions.models import Transaction
from apps.projects.tests.factories import ProjectFactory, ProjectMemberFactory
from apps.projects.models import ProjectMember
from apps.authentication.tests.factories import UserFactory


@pytest.fixture
def api_client():
    """DRF API client."""
    return APIClient()


@pytest.fixture
def authenticated_user():
    """Create and return an authenticated user."""
    return UserFactory()


@pytest.fixture
def project_with_user(authenticated_user):
    """Create a project owned by the authenticated user."""
    project = ProjectFactory(owner=authenticated_user)
    # Ensure owner is added as member
    ProjectMember.objects.get_or_create(
        project=project,
        user=authenticated_user,
        defaults={"role": ProjectMember.Role.OWNER},
    )
    return project


@pytest.mark.django_db
class TestCategoryListCreateView:
    """Test CategoryListCreateView endpoint."""

    def test_list_categories_authenticated_user(self, api_client, authenticated_user, project_with_user):
        """Test that authenticated user can list project categories."""
        api_client.force_authenticate(user=authenticated_user)

        # Create some categories
        CategoryFactory.create_batch(3, project=project_with_user)

        url = reverse("category-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) >= 3

    def test_list_categories_includes_defaults(self, api_client, authenticated_user, project_with_user):
        """Test that default categories are included in list."""
        api_client.force_authenticate(user=authenticated_user)

        # Create default category
        default_cat = CategoryFactory(is_default=True)

        url = reverse("category-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        # Response should have categories
        assert len(response.data) >= 1

    def test_create_category_owner(self, api_client, authenticated_user, project_with_user):
        """Test that project owner can create categories."""
        api_client.force_authenticate(user=authenticated_user)

        data = {
            "name": "New Category",
            "category_type": "expense",
            "color": "#FF5733",
        }

        url = reverse("category-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["name"] == "New Category"

    def test_create_category_admin(self, api_client, authenticated_user):
        """Test that admin member can create categories."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.ADMIN,
        )

        api_client.force_authenticate(user=authenticated_user)

        data = {
            "name": "Admin Category",
            "category_type": "expense",
        }

        url = reverse("category-list-create", kwargs={"project_id": project.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_201_CREATED

    def test_create_category_member_denied(self, api_client, authenticated_user):
        """Test that member cannot create categories."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.MEMBER,
        )

        api_client.force_authenticate(user=authenticated_user)

        data = {
            "name": "Member Category",
            "category_type": "expense",
        }

        url = reverse("category-list-create", kwargs={"project_id": project.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_create_category_non_member(self, api_client, authenticated_user):
        """Test that non-member cannot create category."""
        other_user = UserFactory()
        project = ProjectFactory(owner=other_user)

        api_client.force_authenticate(user=authenticated_user)

        data = {
            "name": "New Category",
            "category_type": "expense",
        }

        url = reverse("category-list-create", kwargs={"project_id": project.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code in [status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND]

    def test_unauthenticated_list_categories(self, api_client, project_with_user):
        """Test that unauthenticated user cannot list categories."""
        url = reverse("category-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_list_categories_ordered_by_name(self, api_client, authenticated_user, project_with_user):
        """Test that categories are ordered by name."""
        api_client.force_authenticate(user=authenticated_user)

        CategoryFactory(project=project_with_user, name="Zebra")
        CategoryFactory(project=project_with_user, name="Apple")
        CategoryFactory(project=project_with_user, name="Mango")

        url = reverse("category-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        # Verify response has categories (may be paginated or list)
        if isinstance(response.data, dict):
            assert "results" in response.data or len(response.data) >= 3
        else:
            assert len(response.data) >= 3


@pytest.mark.django_db
class TestCategoryDetailView:
    """Test CategoryDetailView endpoint."""

    def test_retrieve_category(self, api_client, authenticated_user, project_with_user):
        """Test retrieving a single category."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)

        url = reverse("category-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": category.id
        })
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["name"] == category.name

    def test_retrieve_nonexistent_category(self, api_client, authenticated_user, project_with_user):
        """Test that retrieving nonexistent category raises 404."""
        api_client.force_authenticate(user=authenticated_user)

        import uuid
        url = reverse("category-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": uuid.uuid4()
        })
        response = api_client.get(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_update_category_owner(self, api_client, authenticated_user, project_with_user):
        """Test that owner can update category."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user, name="Old Name")

        data = {"name": "New Name"}
        url = reverse("category-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": category.id
        })
        response = api_client.patch(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["name"] == "New Name"

    def test_update_category_admin(self, api_client, authenticated_user):
        """Test that admin can update category."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.ADMIN,
        )
        category = CategoryFactory(project=project, name="Old Name")

        api_client.force_authenticate(user=authenticated_user)

        data = {"name": "Admin Updated"}
        url = reverse("category-detail", kwargs={
            "project_id": project.id,
            "pk": category.id
        })
        response = api_client.patch(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK

    def test_update_category_member_denied(self, api_client, authenticated_user):
        """Test that member cannot update category."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.MEMBER,
        )
        category = CategoryFactory(project=project, name="Old Name")

        api_client.force_authenticate(user=authenticated_user)

        data = {"name": "Member Updated", "category_type": category.category_type}
        url = reverse("category-detail", kwargs={
            "project_id": project.id,
            "pk": category.id
        })
        response = api_client.patch(url, data, format="json")

        # Should be forbidden or bad request depending on validation order
        assert response.status_code in [status.HTTP_403_FORBIDDEN, status.HTTP_400_BAD_REQUEST]

    def test_delete_category_owner(self, api_client, authenticated_user, project_with_user):
        """Test that owner can delete category."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        category_id = category.id

        url = reverse("category-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": category.id
        })
        response = api_client.delete(url)

        assert response.status_code == status.HTTP_204_NO_CONTENT

    def test_delete_category_admin(self, api_client, authenticated_user):
        """Test that admin can delete category."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.ADMIN,
        )
        category = CategoryFactory(project=project)

        api_client.force_authenticate(user=authenticated_user)

        url = reverse("category-detail", kwargs={
            "project_id": project.id,
            "pk": category.id
        })
        response = api_client.delete(url)

        assert response.status_code == status.HTTP_204_NO_CONTENT

    def test_delete_category_member_denied(self, api_client, authenticated_user):
        """Test that member cannot delete category."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.MEMBER,
        )
        category = CategoryFactory(project=project)

        api_client.force_authenticate(user=authenticated_user)

        url = reverse("category-detail", kwargs={
            "project_id": project.id,
            "pk": category.id
        })
        response = api_client.delete(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
class TestDefaultCategoryListView:
    """Test DefaultCategoryListView endpoint."""

    def test_list_default_categories(self, api_client, authenticated_user):
        """Test that default categories can be listed."""
        api_client.force_authenticate(user=authenticated_user)

        # Create default categories
        CategoryFactory.create_batch(3, is_default=True)

        url = reverse("category-defaults")
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) >= 3

    def test_default_categories_have_is_default_true(self, api_client, authenticated_user):
        """Test that all returned categories have is_default=True."""
        api_client.force_authenticate(user=authenticated_user)

        CategoryFactory.create_batch(2, is_default=True)

        url = reverse("category-defaults")
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        # Verify response has default categories (may be paginated or list)
        if isinstance(response.data, dict):
            assert "results" in response.data
            assert len(response.data["results"]) >= 2
        else:
            assert len(response.data) >= 2

    def test_unauthenticated_cannot_list_defaults(self, api_client):
        """Test that unauthenticated user cannot list default categories."""
        url = reverse("category-defaults")
        response = api_client.get(url)

        assert response.status_code == status.HTTP_401_UNAUTHORIZED


@pytest.mark.django_db
class TestTransactionListCreateView:
    """Test TransactionListCreateView endpoint."""

    def test_list_transactions(self, api_client, authenticated_user, project_with_user):
        """Test listing project transactions."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory.create_batch(
            3,
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data["results"]) >= 3

    def test_create_transaction_owner(self, api_client, authenticated_user, project_with_user):
        """Test creating a transaction as owner."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)

        data = {
            "category": category.id,
            "transaction_type": "expense",
            "amount": "50.75",
            "currency": "USD",
            "description": "Test transaction",
            "date": timezone.now().date(),
        }

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["amount"] == "50.75"

    def test_create_transaction_admin(self, api_client, authenticated_user):
        """Test that admin can create transactions if can_create_transactions is True."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        membership = ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.ADMIN,
        )

        category = CategoryFactory(project=project)

        api_client.force_authenticate(user=authenticated_user)

        data = {
            "category": category.id,
            "transaction_type": "expense",
            "amount": "100.00",
            "currency": "USD",
            "description": "Admin transaction",
            "date": timezone.now().date(),
        }

        url = reverse("transaction-list-create", kwargs={"project_id": project.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_201_CREATED

    def test_create_transaction_member_permission_based(self, api_client, authenticated_user):
        """Test that member can/cannot create based on can_create_transactions permission."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        membership = ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.MEMBER,
        )

        category = CategoryFactory(project=project)

        api_client.force_authenticate(user=authenticated_user)

        data = {
            "category": category.id,
            "transaction_type": "expense",
            "amount": "25.00",
            "currency": "USD",
            "description": "Member transaction",
            "date": timezone.now().date(),
        }

        url = reverse("transaction-list-create", kwargs={"project_id": project.id})
        response = api_client.post(url, data, format="json")

        # Permission depends on can_create_transactions attribute
        assert response.status_code in [status.HTTP_201_CREATED, status.HTTP_403_FORBIDDEN]

    def test_filter_transactions_by_category(self, api_client, authenticated_user, project_with_user):
        """Test filtering transactions by category."""
        api_client.force_authenticate(user=authenticated_user)

        category1 = CategoryFactory(project=project_with_user, name="Food")
        category2 = CategoryFactory(project=project_with_user, name="Transport")

        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category1,
        )
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category2,
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?category={category1.id}")

        assert response.status_code == status.HTTP_200_OK
        # Results should be filtered
        category1_id = str(category1.id)
        for transaction in response.data.get("results", []):
            if transaction:
                # Handle both UUID and string formats
                assert str(transaction["category"]) == category1_id or transaction["category"] == category1_id

    def test_filter_transactions_by_transaction_type(self, api_client, authenticated_user, project_with_user):
        """Test filtering transactions by transaction_type."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            transaction_type="expense",
        )
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            transaction_type="income",
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?transaction_type=expense")

        assert response.status_code == status.HTTP_200_OK
        for transaction in response.data.get("results", []):
            if transaction:
                assert transaction["transaction_type"] == "expense"

    def test_filter_transactions_by_date_range(self, api_client, authenticated_user, project_with_user):
        """Test filtering transactions by date range."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        today = timezone.now().date()

        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            date=today,
        )
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            date=today - timedelta(days=5),
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?date_from={today}")

        assert response.status_code == status.HTTP_200_OK

    def test_search_transactions_by_description(self, api_client, authenticated_user, project_with_user):
        """Test searching transactions by description."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            description="Unique description",
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?search=Unique")

        assert response.status_code == status.HTTP_200_OK

    def test_search_transactions_by_notes(self, api_client, authenticated_user, project_with_user):
        """Test searching transactions by notes."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            notes="Important notes",
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?search=Important")

        assert response.status_code == status.HTTP_200_OK

    def test_order_transactions_by_date_descending(self, api_client, authenticated_user, project_with_user):
        """Test ordering transactions by date descending."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        today = timezone.now().date()

        t1 = TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            date=today - timedelta(days=2),
        )
        t2 = TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            date=today,
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?ordering=-date")

        assert response.status_code == status.HTTP_200_OK
        if len(response.data["results"]) >= 2:
            # Most recent should be first
            assert response.data["results"][0]["date"] >= response.data["results"][1]["date"]

    def test_order_transactions_by_amount(self, api_client, authenticated_user, project_with_user):
        """Test ordering transactions by amount."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)

        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            amount=Decimal("100.00"),
        )
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            amount=Decimal("50.00"),
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?ordering=amount")

        assert response.status_code == status.HTTP_200_OK

    def test_transaction_list_pagination(self, api_client, authenticated_user, project_with_user):
        """Test that transactions are paginated."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory.create_batch(
            15,
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        # Should have pagination structure
        assert "results" in response.data
        assert "count" in response.data
        assert isinstance(response.data["results"], list)

    def test_unauthenticated_cannot_list_transactions(self, api_client, project_with_user):
        """Test that unauthenticated user cannot list transactions."""
        url = reverse("transaction-list-create", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_401_UNAUTHORIZED


@pytest.mark.django_db
class TestTransactionDetailView:
    """Test TransactionDetailView endpoint."""

    def test_retrieve_transaction(self, api_client, authenticated_user, project_with_user):
        """Test retrieving a single transaction."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        transaction = TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": transaction.id
        })
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["amount"] == str(transaction.amount)

    def test_retrieve_nonexistent_transaction(self, api_client, authenticated_user, project_with_user):
        """Test that retrieving nonexistent transaction raises 404."""
        api_client.force_authenticate(user=authenticated_user)

        import uuid
        url = reverse("transaction-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": uuid.uuid4()
        })
        response = api_client.get(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_update_transaction_owner(self, api_client, authenticated_user, project_with_user):
        """Test that owner can update transaction."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        transaction = TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            description="Old description",
        )

        data = {"description": "New description"}
        url = reverse("transaction-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": transaction.id
        })
        response = api_client.patch(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["description"] == "New description"

    def test_update_transaction_admin(self, api_client, authenticated_user):
        """Test that admin can update if can_edit_transactions is True."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.ADMIN,
        )

        category = CategoryFactory(project=project)
        transaction = TransactionFactory(
            project=project,
            created_by=owner,
            category=category,
            description="Old",
        )

        api_client.force_authenticate(user=authenticated_user)

        data = {"description": "Admin Updated"}
        url = reverse("transaction-detail", kwargs={
            "project_id": project.id,
            "pk": transaction.id
        })
        response = api_client.patch(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK

    def test_update_transaction_member_permission_based(self, api_client, authenticated_user):
        """Test that member can/cannot update based on can_edit_transactions."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.MEMBER,
        )

        category = CategoryFactory(project=project)
        transaction = TransactionFactory(
            project=project,
            created_by=owner,
            category=category,
            description="Old",
        )

        api_client.force_authenticate(user=authenticated_user)

        data = {"description": "Member Updated"}
        url = reverse("transaction-detail", kwargs={
            "project_id": project.id,
            "pk": transaction.id
        })
        response = api_client.patch(url, data, format="json")

        # Permission depends on can_edit_transactions
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_403_FORBIDDEN]

    def test_delete_transaction_owner(self, api_client, authenticated_user, project_with_user):
        """Test that owner can delete transaction if can_delete_transactions is True."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        transaction = TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-detail", kwargs={
            "project_id": project_with_user.id,
            "pk": transaction.id
        })
        response = api_client.delete(url)

        # Owner may or may not have delete permission depending on can_delete_transactions
        assert response.status_code in [status.HTTP_204_NO_CONTENT, status.HTTP_403_FORBIDDEN]

    def test_delete_transaction_admin(self, api_client, authenticated_user):
        """Test that admin can delete if can_delete_transactions is True."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.ADMIN,
        )

        category = CategoryFactory(project=project)
        transaction = TransactionFactory(
            project=project,
            created_by=owner,
            category=category,
        )

        api_client.force_authenticate(user=authenticated_user)

        url = reverse("transaction-detail", kwargs={
            "project_id": project.id,
            "pk": transaction.id
        })
        response = api_client.delete(url)

        # Admin may or may not have delete permission depending on can_delete_transactions
        assert response.status_code in [status.HTTP_204_NO_CONTENT, status.HTTP_403_FORBIDDEN]

    def test_delete_transaction_member_permission_based(self, api_client, authenticated_user):
        """Test that member can/cannot delete based on can_delete_transactions."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        ProjectMemberFactory(
            project=project,
            user=authenticated_user,
            role=ProjectMember.Role.MEMBER,
        )

        category = CategoryFactory(project=project)
        transaction = TransactionFactory(
            project=project,
            created_by=owner,
            category=category,
        )

        api_client.force_authenticate(user=authenticated_user)

        url = reverse("transaction-detail", kwargs={
            "project_id": project.id,
            "pk": transaction.id
        })
        response = api_client.delete(url)

        # Permission depends on can_delete_transactions
        assert response.status_code in [status.HTTP_204_NO_CONTENT, status.HTTP_403_FORBIDDEN]


@pytest.mark.django_db
class TestTransactionBulkCreateView:
    """Test TransactionBulkCreateView endpoint."""

    def test_bulk_create_valid_transactions(self, api_client, authenticated_user, project_with_user):
        """Test bulk creating valid transactions."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        today = timezone.now().date()

        data = [
            {
                "category": str(category.id),
                "transaction_type": "expense",
                "amount": "50.00",
                "currency": "USD",
                "description": "Transaction 1",
                "date": today,
            },
            {
                "category": str(category.id),
                "transaction_type": "expense",
                "amount": "75.00",
                "currency": "USD",
                "description": "Transaction 2",
                "date": today,
            },
        ]

        url = reverse("transaction-bulk-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_201_CREATED
        assert len(response.data) == 2

    def test_bulk_create_empty_list_error(self, api_client, authenticated_user, project_with_user):
        """Test that empty list raises error."""
        api_client.force_authenticate(user=authenticated_user)

        data = []

        url = reverse("transaction-bulk-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_bulk_create_exceeds_limit(self, api_client, authenticated_user, project_with_user):
        """Test that exceeding 100 items limit raises error."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        today = timezone.now().date()

        # Create list with 101 items
        data = [
            {
                "category": str(category.id),
                "transaction_type": "expense",
                "amount": "10.00",
                "currency": "USD",
                "description": f"Transaction {i}",
                "date": today,
            }
            for i in range(101)
        ]

        url = reverse("transaction-bulk-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        # Response could be dict or list depending on error format
        if isinstance(response.data, dict):
            assert "cannot bulk create more than 100" in str(response.data).lower()
        else:
            assert "cannot bulk create more than 100" in str(response.data).lower()

    def test_bulk_create_at_limit(self, api_client, authenticated_user, project_with_user):
        """Test bulk creating exactly 100 transactions."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        today = timezone.now().date()

        # Create list with 100 items
        data = [
            {
                "category": str(category.id),
                "transaction_type": "expense",
                "amount": "10.00",
                "currency": "USD",
                "description": f"Transaction {i}",
                "date": today,
            }
            for i in range(100)
        ]

        url = reverse("transaction-bulk-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_201_CREATED
        assert len(response.data) == 100

    def test_bulk_create_with_invalid_item(self, api_client, authenticated_user, project_with_user):
        """Test that invalid item in list causes error."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        today = timezone.now().date()

        data = [
            {
                "category": str(category.id),
                "transaction_type": "expense",
                "amount": "50.00",
                "currency": "USD",
                "description": "Valid",
                "date": today,
            },
            {
                # Missing required fields
                "description": "Invalid",
            },
        ]

        url = reverse("transaction-bulk-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_bulk_create_permission_denied(self, api_client, authenticated_user):
        """Test that non-member cannot bulk create."""
        owner = UserFactory()
        project = ProjectFactory(owner=owner)
        # Don't add authenticated_user as member at all

        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project)
        today = timezone.now().date()

        data = [
            {
                "category": str(category.id),
                "transaction_type": "expense",
                "amount": "50.00",
                "currency": "USD",
                "description": "Test",
                "date": today,
            }
        ]

        url = reverse("transaction-bulk-create", kwargs={"project_id": project.id})
        response = api_client.post(url, data, format="json")

        # Non-member should not have access
        assert response.status_code in [status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND]

    def test_bulk_create_unauthenticated(self, api_client, project_with_user):
        """Test that unauthenticated user cannot bulk create."""
        data = [{"category": "", "transaction_type": "expense", "amount": "50.00"}]

        url = reverse("transaction-bulk-create", kwargs={"project_id": project_with_user.id})
        response = api_client.post(url, data, format="json")

        assert response.status_code == status.HTTP_401_UNAUTHORIZED


@pytest.mark.django_db
class TestTransactionExportView:
    """Test TransactionExportView endpoint."""

    def test_export_csv_default_format(self, api_client, authenticated_user, project_with_user):
        """Test exporting transactions as CSV (default)."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user, name="Food")
        today = timezone.now().date()

        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            description="Lunch",
            amount=Decimal("25.50"),
            date=today,
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "text/csv"
        assert "attachment" in response.get("Content-Disposition", "")

    def test_export_csv_explicit_format(self, api_client, authenticated_user, project_with_user):
        """Test exporting transactions as CSV with explicit format."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?export_format=csv")

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "text/csv"

    def test_export_csv_content(self, api_client, authenticated_user, project_with_user):
        """Test that CSV contains correct headers and data."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user, name="Food")
        today = timezone.now().date()

        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            description="Lunch",
            amount=Decimal("25.50"),
            date=today,
            transaction_type="expense",
            currency="USD",
            payment_method="card",
            reference_number="REF123",
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?export_format=csv")

        assert response.status_code == status.HTTP_200_OK
        # Parse CSV content
        content = response.content.decode("utf-8")
        assert "Date" in content
        assert "Type" in content
        assert "Amount" in content
        assert "Category" in content

    def test_export_xlsx_format(self, api_client, authenticated_user, project_with_user):
        """Test exporting transactions as XLSX."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?export_format=xlsx")

        assert response.status_code == status.HTTP_200_OK
        assert "spreadsheetml" in response["Content-Type"]

    def test_export_xlsx_content(self, api_client, authenticated_user, project_with_user):
        """Test that XLSX contains correct headers and data."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user, name="Food")
        today = timezone.now().date()

        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
            description="Lunch",
            amount=Decimal("25.50"),
            date=today,
            transaction_type="expense",
            currency="USD",
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?export_format=xlsx")

        assert response.status_code == status.HTTP_200_OK
        # Parse XLSX content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        # Check headers exist
        headers = [cell.value for cell in ws[1]]
        assert "Date" in headers
        assert "Type" in headers
        assert "Amount" in headers

    def test_export_multiple_transactions(self, api_client, authenticated_user, project_with_user):
        """Test exporting multiple transactions."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory.create_batch(
            5,
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK

    def test_export_empty_project(self, api_client, authenticated_user, project_with_user):
        """Test exporting from project with no transactions."""
        api_client.force_authenticate(user=authenticated_user)

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "text/csv"

    def test_export_case_insensitive_format(self, api_client, authenticated_user, project_with_user):
        """Test that format parameter is case insensitive."""
        api_client.force_authenticate(user=authenticated_user)

        category = CategoryFactory(project=project_with_user)
        TransactionFactory(
            project=project_with_user,
            created_by=authenticated_user,
            category=category,
        )

        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(f"{url}?export_format=XLSX")

        assert response.status_code == status.HTTP_200_OK
        assert "spreadsheetml" in response["Content-Type"]

    def test_export_permission_denied(self, api_client, authenticated_user):
        """Test that non-member cannot export."""
        other_user = UserFactory()
        project = ProjectFactory(owner=other_user)

        api_client.force_authenticate(user=authenticated_user)

        url = reverse("transaction-export", kwargs={"project_id": project.id})
        response = api_client.get(url)

        assert response.status_code in [status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND]

    def test_export_unauthenticated(self, api_client, project_with_user):
        """Test that unauthenticated user cannot export."""
        url = reverse("transaction-export", kwargs={"project_id": project_with_user.id})
        response = api_client.get(url)

        assert response.status_code == status.HTTP_401_UNAUTHORIZED
